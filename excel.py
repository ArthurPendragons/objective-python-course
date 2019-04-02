import openpyxl
import os

workbook = openpyxl.load_workbook('email.xlsx')
sheet = workbook['Sheet1']

#cell = sheet['A2']
#print(cell.value.split()[0])
#print(sheet.cell(row=1,column=2))

def getName(number):
	return sheet.cell(row=number, column=1).value.split()[0]

def getEmail(number):
		return sheet.cell(row=number, column=2).value

for x in range(2,9):
	currentName = getName(x)
	email = getEmail(x)
	print('Hi {}, your email is {}'.format(currentName,email))